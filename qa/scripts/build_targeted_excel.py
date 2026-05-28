"""Build timestamped Excel + DECISION_ANALYSIS.md from targeted JSONLs."""
from __future__ import annotations
import json, os, sys
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule

base = REPO_ROOT / "qa" / "runs" / "20260512-targeted"
methods = ("normal", "mmr", "fuzzy", "embedding")
data = {}
for m in methods:
    rows = [json.loads(l) for l in (base / f"targeted_{m}.jsonl").read_text(encoding="utf-8").splitlines() if l.strip()]
    data[m] = {r["case"]: r for r in rows}

ts = datetime.now()
ts_fname = ts.strftime("%Y-%m-%d_%H-%M")
out_xlsx = REPO_ROOT / "qa" / "test-cases" / f"master_test_cases_{ts_fname}.xlsx"
out_xlsx.parent.mkdir(parents=True, exist_ok=True)

HFILL = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN = Border(*([Side(style="thin", color="BFBFBF")] * 4))
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
GRAY = PatternFill("solid", fgColor="D9D9D9")
WIN = PatternFill("solid", fgColor="9BC2E6")


def style_header(ws, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HFILL
        cell.font = HFONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")


def style_body(ws, last, n):
    for r in range(2, last + 1):
        for c in range(1, n + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            cell.alignment = WRAP


def widths(ws, m):
    for k, v in m.items():
        ws.column_dimensions[k].width = v


def cf_status(ws, col, last):
    rng = f"{col}2:{col}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=GREEN))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=RED))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"SKIP"'], fill=GRAY))


# Aggregate
total_dur_ms = 0
for m in methods:
    t1 = data[m]["T1"]
    for t in t1["per_turn"]:
        total_dur_ms += t.get("wallclock_ms") or 0
    total_dur_ms += data[m]["T2"].get("wallclock_ms") or 0
    total_dur_ms += data[m]["T3"].get("wallclock_ms_recap") or 0
total_dur_min = total_dur_ms / 1000 / 60


def pass_count(m):
    n = 0
    t1 = data[m]["T1"]
    if m == "normal":
        n += 1 if t1["rc_count_final"] == 0 else 0
    else:
        n += 1 if (t1["rc_growth_turn1_to_3"] >= 4 and t1["rc_count_final"] >= 4) else 0
    t2 = data[m]["T2"]
    n += 1 if t2["retrieval_method"] == m else 0
    t3 = data[m]["T3"]
    if m == "normal":
        n += 1 if (t3["pre_turn_rc_count"] == 0 and not t3["rc_grew_on_recap"]) else 0
    else:
        n += 1 if (t3["pre_turn_rc_count"] >= 1 and not t3["rc_grew_on_recap"]) else 0
    return n


wb = Workbook()
# Sheet 1: Summary
s = wb.active
s.title = "Summary"
rows = [
    ("Run started", "2026-05-12 (post-KB-rebuild)"),
    ("Completed at", ts.strftime("%Y-%m-%d %H:%M:%S")),
    ("Total wall duration (min)", f"{total_dur_min:.1f}"),
    ("Total user turns sent", 4 * 6),
    ("Total mongo audit rows captured", 24),
    ("Target URL", "http://localhost:2305"),
    ("KB checksum (post-rebuild)", "meta_id 6a0222ece6aaaa5fd1f36693 (legacy single-collection, 499 docs)"),
    ("Embedding provider", "HuggingFace all-MiniLM-L6-v2 (local, no OpenAI cost)"),
    ("REDUNDANCY_METHOD path verified", "incontext_service_validation (FAQ-RAG completion)"),
    ("", ""),
    ("Method", "Test PASS / Total"),
]
for m in methods:
    rows.append((m, f"{pass_count(m)} / 3"))
rows.append(("", ""))
rows.append(("Top-line verdict",
    "MMR is the most effective anti-repetition method: rc_count_final=11 (vs 10 for fuzzy/embedding), "
    "with no latency penalty on FAQ-RAG path. All 3 non-normal methods correctly handle recap bypass."))
for i, (k, v) in enumerate(rows, start=1):
    s.cell(row=i, column=1, value=k)
    s.cell(row=i, column=2, value=v)
widths(s, {"A": 32, "B": 110})

# Sheet 2: Test Cases
s_tc = wb.create_sheet("Test Cases")
s_tc.append(["Case ID", "Goal", "Setup", "Metric", "Expected", "API budget"])
s_tc.append(["T1", "Multi-turn anti-repetition",
             "3 turns same topic in fresh session",
             "recent_chunk_ids growth, new chunks per turn",
             "Normal: rc stays 0. Non-normal: rc grows by ~4 IDs per turn",
             "12 user msgs"])
s_tc.append(["T2", "Within-turn diversity",
             "1 catalog query in fresh session",
             "distinct_qstems, distinct_services in Context",
             "All methods: q_stems>=3. MMR/Embedding may edge on redundant queries",
             "4 user msgs"])
s_tc.append(["T3", "Recap bypass",
             "2 turns: pre + recap phrase",
             "rc growth on recap turn",
             "Non-normal: pre populates rc, recap turn rc unchanged (grew=False). Normal: rc stays 0 both turns",
             "8 user msgs"])
last = s_tc.max_row
style_header(s_tc, 6)
style_body(s_tc, last, 6)
widths(s_tc, {"A": 8, "B": 28, "C": 30, "D": 36, "E": 60, "F": 14})
s_tc.freeze_panes = "B2"

# Sheet 3: Results
s_r = wb.create_sheet("Results")
s_r.append(["Case ID", "Method", "Metric Name", "Expected", "Actual", "Status", "Delta vs Normal", "Notes"])
for cid in ("T1", "T2", "T3"):
    for m in methods:
        r = data[m][cid]
        if cid == "T1":
            metric = "rc_count_final / growth_T1_to_T3"
            actual = f"{r['rc_count_final']} / {r['rc_growth_turn1_to_3']}"
            if m == "normal":
                expected = "0 / 0 (no tracking)"
                status = "PASS" if r["rc_count_final"] == 0 else "FAIL"
                delta = "baseline"
            else:
                expected = ">=4 / >=4 (filter actively writing)"
                status = "PASS" if (r["rc_growth_turn1_to_3"] >= 4 and r["rc_count_final"] >= 4) else "FAIL"
                delta = f"+{r['rc_count_final']}"
            notes = f"per_turn_rc=[{r['per_turn'][0]['rc_count']},{r['per_turn'][1]['rc_count']},{r['per_turn'][2]['rc_count']}]"
        elif cid == "T2":
            metric = "distinct_qstems / services"
            actual = f"{r['distinct_qstems']} / {r['distinct_services']}"
            expected = ">=3 / >=2"
            status = "PASS" if (r["distinct_qstems"] >= 3 and r["distinct_services"] >= 2) else "FAIL"
            nor_q = data["normal"]["T2"]["distinct_qstems"]
            delta = f"{r['distinct_qstems'] - nor_q:+d}" if m != "normal" else "baseline"
            notes = f"route={r['route']}, rm={r['retrieval_method']}"
        else:
            metric = "rc_grew_on_recap"
            actual = f"pre={r['pre_turn_rc_count']}, recap={r['recap_turn_rc_count']}, grew={r['rc_grew_on_recap']}"
            if m == "normal":
                expected = "rc stays 0 both turns"
                status = "PASS" if (r["pre_turn_rc_count"] == 0 and not r["rc_grew_on_recap"]) else "FAIL"
                delta = "baseline"
            else:
                expected = "pre populated (>=1), recap unchanged"
                status = "PASS" if (r["pre_turn_rc_count"] >= 1 and not r["rc_grew_on_recap"]) else "FAIL"
                delta = f"pre+{r['pre_turn_rc_count']}"
            notes = f"rm_pre={r['rm_pre']}, rm_recap={r['rm_recap']}"
        s_r.append([cid, m, metric, expected, actual, status, delta, notes])
last = s_r.max_row
style_header(s_r, 8)
style_body(s_r, last, 8)
cf_status(s_r, "F", last)
widths(s_r, {"A": 8, "B": 12, "C": 28, "D": 36, "E": 40, "F": 10, "G": 14, "H": 38})
s_r.freeze_panes = "C2"

# Sheet 4: Comparison Matrix
s_c = wb.create_sheet("Comparison Matrix")
s_c.append(["Test", "Metric", "normal", "mmr", "fuzzy", "embedding", "Winner"])
matrix_rows = []
# T1
t1n = data["normal"]["T1"]
t1m = data["mmr"]["T1"]
t1f = data["fuzzy"]["T1"]
t1e = data["embedding"]["T1"]
matrix_rows.append(("T1", "rc_count_final (after 3 turns)", t1n["rc_count_final"], t1m["rc_count_final"], t1f["rc_count_final"], t1e["rc_count_final"], "max"))
matrix_rows.append(("T1", "rc_growth_T1_to_T3", t1n["rc_growth_turn1_to_3"], t1m["rc_growth_turn1_to_3"], t1f["rc_growth_turn1_to_3"], t1e["rc_growth_turn1_to_3"], "max"))
matrix_rows.append(("T1", "new chunks in T2", t1n["rc_new_in_turn2"], t1m["rc_new_in_turn2"], t1f["rc_new_in_turn2"], t1e["rc_new_in_turn2"], "max"))
matrix_rows.append(("T1", "new chunks in T3", t1n["rc_new_in_turn3"], t1m["rc_new_in_turn3"], t1f["rc_new_in_turn3"], t1e["rc_new_in_turn3"], "max"))
matrix_rows.append(("T1", "mean turn latency (ms)",
    int(sum(t["wallclock_ms"] or 0 for t in t1n["per_turn"]) / 3),
    int(sum(t["wallclock_ms"] or 0 for t in t1m["per_turn"]) / 3),
    int(sum(t["wallclock_ms"] or 0 for t in t1f["per_turn"]) / 3),
    int(sum(t["wallclock_ms"] or 0 for t in t1e["per_turn"]) / 3),
    "min"))
# T2
t2n = data["normal"]["T2"]
t2m = data["mmr"]["T2"]
t2f = data["fuzzy"]["T2"]
t2e = data["embedding"]["T2"]
matrix_rows.append(("T2", "distinct_qstems", t2n["distinct_qstems"], t2m["distinct_qstems"], t2f["distinct_qstems"], t2e["distinct_qstems"], "max"))
matrix_rows.append(("T2", "distinct_services", t2n["distinct_services"], t2m["distinct_services"], t2f["distinct_services"], t2e["distinct_services"], "max"))
matrix_rows.append(("T2", "latency (ms)", t2n["wallclock_ms"], t2m["wallclock_ms"], t2f["wallclock_ms"], t2e["wallclock_ms"], "min"))
# T3
t3n = data["normal"]["T3"]
t3m = data["mmr"]["T3"]
t3f = data["fuzzy"]["T3"]
t3e = data["embedding"]["T3"]
matrix_rows.append(("T3", "pre_turn_rc", t3n["pre_turn_rc_count"], t3m["pre_turn_rc_count"], t3f["pre_turn_rc_count"], t3e["pre_turn_rc_count"], "max"))
matrix_rows.append(("T3", "recap_turn_rc", t3n["recap_turn_rc_count"], t3m["recap_turn_rc_count"], t3f["recap_turn_rc_count"], t3e["recap_turn_rc_count"], "max"))
matrix_rows.append(("T3", "rc_grew_on_recap (False=correct)", str(t3n["rc_grew_on_recap"]), str(t3m["rc_grew_on_recap"]), str(t3f["rc_grew_on_recap"]), str(t3e["rc_grew_on_recap"]), "all-false"))

for test, metric, n, mmr, fz, em, mode in matrix_rows:
    if mode == "all-false":
        win = "tie (all correct)" if {n, mmr, fz, em} == {"False"} else "split"
    else:
        vals = [n, mmr, fz, em]
        try:
            num_vals = [int(v) for v in vals]
            if len(set(num_vals)) == 1:
                win = "tie"
            else:
                target = min(num_vals) if mode == "min" else max(num_vals)
                idx = num_vals.index(target)
                win = ("normal", "mmr", "fuzzy", "embedding")[idx]
        except Exception:
            win = "n/a"
    s_c.append([test, metric, n, mmr, fz, em, win])
last = s_c.max_row
style_header(s_c, 7)
style_body(s_c, last, 7)
for r in range(2, last + 1):
    win = s_c.cell(row=r, column=7).value
    if win in ("normal", "mmr", "fuzzy", "embedding"):
        col_map = {"normal": 3, "mmr": 4, "fuzzy": 5, "embedding": 6}
        s_c.cell(row=r, column=col_map[win]).fill = WIN
widths(s_c, {"A": 8, "B": 38, "C": 12, "D": 12, "E": 12, "F": 14, "G": 22})
s_c.freeze_panes = "C2"

# Sheet 5: Decision
s_d = wb.create_sheet("Decision")
decision = [
    "Decision Summary",
    "",
    "1. Anti-redundancy core feature (T1):",
    "   WINNER: mmr",
    "   Evidence: rc_count_final=11 vs fuzzy/embedding=10 (+10%); rc_growth=7 vs 6 (+17%).",
    "   MMR is most aggressive in bringing FRESH chunks across turns.",
    "",
    "2. Within-turn diversity (T2):",
    "   WINNER: tie (all 4 methods identical for this catalog query)",
    "   Evidence: distinct_qstems=4, distinct_services=2 for all methods.",
    "   No within-turn advantage on this query. Earlier Module B (B01 whistleblowing) showed mmr/embedding +1 q_stem on redundant queries — not reproducible here because catalog query has naturally diverse Q-stems.",
    "",
    "3. Recap bypass correctness (T3):",
    "   WINNER: tie (mmr=fuzzy=embedding all correct)",
    "   Evidence: pre_rc=4, recap_rc=4, grew=False for all 3 non-normal methods. Normal correctly has rc=0 throughout.",
    "",
    "4. Recommended default: mmr",
    "   Justification: best anti-repetition (rc_count +1 over fuzzy/embedding, growth +1) + correct recap bypass + parity on within-turn diversity + ZERO latency penalty on FAQ-RAG path (mean turn -9% vs normal, likely due to smaller post-filter prompt).",
    "",
    "5. CAVEAT (critical for interpretation):",
    "   These results valid ONLY for queries routing to `incontext_service_validation` (FAQ-RAG completion path).",
    "   Queries routing to `sa_compose` (SA-handoff path) — most domain questions like 'apa itu whistleblowing' — bypass writeback per Issue #1 (Task 20 SA-handoff observability gap, DEFERRED).",
    "   Real production traffic mix should be measured before flipping default to mmr in production.",
]
for r in decision:
    s_d.append([r])
widths(s_d, {"A": 130})

wb.save(out_xlsx)
print(f"WROTE: {out_xlsx}")

# Save aggregated metrics
metrics_json = base / "aggregated_metrics.json"
metrics_json.write_text(json.dumps({m: data[m] for m in methods}, indent=2, default=str), encoding="utf-8")
print(f"METRICS: {metrics_json}")
print(f"TIMESTAMP: {ts_fname}")
