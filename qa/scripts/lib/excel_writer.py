"""Aggregate per-method JSONL results into a single Excel workbook.

Sheet layout (per spec):
1. Summary
2. Test Cases
3. Results — Normal
4. Results — MMR
5. Results — Fuzzy
6. Results — Embedding
7. Comparison
8. Performance
9. Verdict
"""
from __future__ import annotations
import json
from pathlib import Path
from typing import Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN_BORDER = Border(*([Side(style="thin", color="BFBFBF")] * 4))
WRAP = Alignment(wrap_text=True, vertical="top")

STATUS_GREEN = PatternFill("solid", fgColor="C6EFCE")
STATUS_RED = PatternFill("solid", fgColor="FFC7CE")
STATUS_YELLOW = PatternFill("solid", fgColor="FFEB9C")


def _style_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")


def _style_body(ws, n_rows: int, n_cols: int) -> None:
    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = WRAP


def _set_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _apply_status_cf(ws, status_col_letter: str, last_row: int) -> None:
    if last_row < 2:
        return  # No data rows, skip conditional formatting
    rng = f"{status_col_letter}2:{status_col_letter}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=STATUS_GREEN))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=STATUS_RED))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PARTIAL"'], fill=STATUS_YELLOW))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"ERROR"'], fill=STATUS_RED))


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    out = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            out.append(json.loads(line))
    return out


def write_workbook(
    *,
    run_dir: Path,
    output_xlsx: Path,
    kb_meta: dict[str, Any],
    target_url: str,
) -> None:
    """Aggregate runs/<timestamp>/method_*.jsonl into a single Excel workbook."""
    methods = ["normal", "mmr", "fuzzy", "embedding"]
    by_method: dict[str, list[dict[str, Any]]] = {
        m: _read_jsonl(run_dir / f"method_{m}.jsonl") for m in methods
    }

    wb = Workbook()
    # Default sheet → repurposed as "Summary"
    s_sum = wb.active
    s_sum.title = "Summary"

    # ----- Summary -----
    rows: list[tuple[str, Any]] = [
        ("Run started", run_dir.name),
        ("Target URL", target_url),
        ("KB checksum", kb_meta.get("checksum", "")),
        ("KB doc_count", kb_meta.get("doc_count", "")),
        ("KB built_at", kb_meta.get("built_at", "")),
        ("Embedding label", kb_meta.get("embedding_label", "")),
        ("", ""),
        ("Method", "Total / Pass / Fail / Partial / Error"),
    ]
    for m in methods:
        results = by_method[m]
        n = len(results)
        passed = sum(1 for r in results if r.get("status") == "PASS")
        failed = sum(1 for r in results if r.get("status") == "FAIL")
        partial = sum(1 for r in results if r.get("status") == "PARTIAL")
        err = sum(1 for r in results if r.get("status") == "ERROR")
        rows.append((m, f"{n} / {passed} / {failed} / {partial} / {err}"))
    for i, (k, v) in enumerate(rows, start=1):
        s_sum.cell(row=i, column=1, value=k)
        s_sum.cell(row=i, column=2, value=v)
    s_sum.column_dimensions["A"].width = 22
    s_sum.column_dimensions["B"].width = 60

    # ----- Results per method (sheets 3-6) -----
    result_cols = [
        "case_id", "module", "type", "test_case", "status", "expected", "actual",
        "delta_vs_normal", "wallclock_ms", "session_id", "notes",
    ]
    for m in methods:
        ws = wb.create_sheet(f"Results — {m.title()}")
        ws.append([c.replace("_", " ").title() for c in result_cols])
        for row in by_method[m]:
            ws.append([row.get(c, "") for c in result_cols])
        last = ws.max_row
        n_cols = len(result_cols)
        _style_header(ws, n_cols)
        _style_body(ws, last, n_cols)
        _set_widths(ws, {
            "A": 12, "B": 12, "C": 14, "D": 40, "E": 10, "F": 50, "G": 50,
            "H": 16, "I": 14, "J": 38, "K": 50,
        })
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        _apply_status_cf(ws, "E", last)

    # ----- Test Cases sheet (catalog of all cases from normal run) -----
    s_tc = wb.create_sheet("Test Cases", index=1)
    tc_cols = ["case_id", "module", "type", "test_case", "expected", "notes"]
    s_tc.append([c.replace("_", " ").title() for c in tc_cols])
    for row in by_method["normal"]:
        s_tc.append([row.get(c, "") for c in tc_cols])
    last = s_tc.max_row
    n_cols = len(tc_cols)
    _style_header(s_tc, n_cols)
    _style_body(s_tc, last, n_cols)
    _set_widths(s_tc, {"A": 12, "B": 12, "C": 14, "D": 40, "E": 50, "F": 50})
    s_tc.freeze_panes = "B2"
    s_tc.auto_filter.ref = s_tc.dimensions

    # ----- Comparison sheet -----
    s_cmp = wb.create_sheet("Comparison")
    cmp_cols = [
        "case_id", "module", "test_case",
        "normal_status", "mmr_status", "fuzzy_status", "embedding_status",
        "normal_metric", "mmr_metric", "fuzzy_metric", "embedding_metric",
        "delta_mmr", "delta_fuzzy", "delta_embedding",
    ]
    s_cmp.append([c.replace("_", " ").title() for c in cmp_cols])
    # Index normal by case_id
    norm_by_id = {r["case_id"]: r for r in by_method["normal"]}
    seen_ids: list[str] = []
    for r in by_method["normal"]:
        cid = r["case_id"]
        if cid in seen_ids:
            continue
        seen_ids.append(cid)
        row_out = {
            "case_id": cid,
            "module": r.get("module", ""),
            "test_case": r.get("test_case", ""),
            "normal_status": r.get("status", ""),
            "normal_metric": r.get("metric", ""),
        }
        for m in ("mmr", "fuzzy", "embedding"):
            mr = next((x for x in by_method[m] if x["case_id"] == cid), {})
            row_out[f"{m}_status"] = mr.get("status", "")
            row_out[f"{m}_metric"] = mr.get("metric", "")
            row_out[f"delta_{m}"] = mr.get("delta_vs_normal", "")
        s_cmp.append([row_out.get(c, "") for c in cmp_cols])
    last = s_cmp.max_row
    n_cols = len(cmp_cols)
    _style_header(s_cmp, n_cols)
    _style_body(s_cmp, last, n_cols)
    s_cmp.freeze_panes = "B2"
    s_cmp.auto_filter.ref = s_cmp.dimensions

    # ----- Performance sheet -----
    s_perf = wb.create_sheet("Performance")
    s_perf.append(["Method", "Sample Count", "p50 (ms)", "p95 (ms)", "p99 (ms)", "Max (ms)"])
    for m in methods:
        durs = sorted(int(r.get("wallclock_ms") or 0) for r in by_method[m] if r.get("wallclock_ms"))
        n = len(durs)
        if n == 0:
            s_perf.append([m, 0, "", "", "", ""])
            continue
        def pct(p: float) -> int:
            idx = min(n - 1, int(round((n - 1) * p)))
            return durs[idx]
        s_perf.append([m, n, pct(0.50), pct(0.95), pct(0.99), max(durs)])
    _style_header(s_perf, 6)
    _style_body(s_perf, s_perf.max_row, 6)
    _set_widths(s_perf, {"A": 14, "B": 14, "C": 12, "D": 12, "E": 12, "F": 12})

    # ----- Verdict sheet -----
    s_v = wb.create_sheet("Verdict")
    s_v.append(["Method", "Verdict", "Recommendation"])
    norm_n = len(by_method["normal"])
    norm_pass = sum(1 for r in by_method["normal"] if r.get("status") == "PASS")
    for m in ("mmr", "fuzzy", "embedding"):
        m_results = by_method[m]
        m_n = len(m_results)
        if m_n == 0:
            verdict = "NOT RUN"
            reco = "Method not executed in this run."
        elif norm_n == 0:
            verdict = "NO BASELINE"
            reco = "Normal baseline missing — comparison unavailable."
        else:
            pass_count = sum(1 for r in m_results if r.get("status") == "PASS")
            fail_count = sum(1 for r in m_results if r.get("status") == "FAIL")
            if pass_count == 0 and fail_count > 0:
                verdict = "FAILED"
                reco = f"All {fail_count} tests failed. Investigate before considering deployment."
            else:
                diff = pass_count - norm_pass
                if diff > 0:
                    verdict = "IMPROVED"
                    reco = f"+{diff} more tests pass than normal. Consider as candidate default."
                elif diff < 0:
                    verdict = "REGRESSED"
                    reco = f"{diff} regression vs normal. Investigate failed cases."
                else:
                    verdict = "NO CHANGE"
                    reco = "Behavior parity with normal. Investigate diversity metrics in Comparison sheet."
        s_v.append([m, verdict, reco])
    _style_header(s_v, 3)
    _style_body(s_v, s_v.max_row, 3)
    _set_widths(s_v, {"A": 14, "B": 14, "C": 70})

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
